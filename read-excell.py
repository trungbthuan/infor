import openpyxl
import sqlite3
from datetime import datetime
from openpyxl.utils import datetime as xl_datetime

def parse_date(value):
    if value is None:
        return None

    if isinstance(value, datetime):
        return value.date()

    if isinstance(value, (float, int)):
        try:
            return xl_datetime.from_excel(value).date()
        except:
            return None

    if isinstance(value, str):
        text = value.strip()
        if text == "":
            return None
        try:
            return datetime.strptime(text, "%d/%m/%Y").date()
        except:
            print("❗ Ngày không hợp lệ:", text)
            return None

    return None

df = openpyxl.load_workbook("profiles.xlsx")
sheet = df.active

conn = sqlite3.connect('db.sqlite3')
cursor = conn.cursor()

row_values = []

# -------- CÁCH 2 — In từng ô theo dạng đẹp hơn ----------------
# for row in sheet.iter_rows(values_only=True):
#     for cell in row:
#         print(cell, end=" | ")
#     print() 

# -------- CÁCH 3 — In theo chỉ số dòng (row number) -----------
# for r in range(2, sheet.max_row + 1):
#     for c in range(1, sheet.max_column + 1):
#         row_values.append(sheet.cell(row=r, column=c).value)
#     print(row_values)

# ----------------------------------------------------
for r in range(2, sheet.max_row + 1):
    try:
        birthday = parse_date(sheet.cell(row=r, column=7).value)
        recruitment_day = parse_date(sheet.cell(row=r, column=4).value)

        query = """
            INSERT INTO api_profiles 
            (id, full_name, birthday, sex, birth_place, nation, recruitment_day, job_title, department) 
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """
        val = (
            sheet.cell(row=r, column=2).value,   # id
            sheet.cell(row=r, column=3).value,   # full_name
            birthday,                            # birthday
            sheet.cell(row=r, column=1).value,   # sex
            sheet.cell(row=r, column=8).value,   # birth_place
            sheet.cell(row=r, column=9).value,   # nation
            recruitment_day,                     # recruitment_day
            sheet.cell(row=r, column=5).value,   # job_title
            sheet.cell(row=r, column=6).value    # department
        )
        cursor.execute(query, val) 
    except sqlite3.Error as error:
        print('Error occurred - ', error)
    
conn.commit()        
conn.close()
print('SQLite Connection closed')


# for row in range(1, sheet_obj.max_row + 1):
#     for col in sheet_obj.iter_cols(1, sheet_obj.max_column-1):
#         print(col[0].value, ': ', col[row+1].value)